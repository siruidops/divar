#!/usr/bin/env python3

import requests
from bs4 import BeautifulSoup
import time
import datetime as dt
import sys
import os
import threading
from stem import Signal
from stem.control import Controller
from openpyxl import load_workbook
from openpyxl import Workbook
import random
import datetime

city_list = ['http://divar.ir/s/tehran']
id_list = []

lock1 = threading.Lock()
lock2 = threading.Lock()
lock3 = threading.Lock()

datenow = dt.datetime.now()
year = datenow.year; month = datenow.month; day = datenow.day
timenow = "{}_{}_{}".format(year,month,day)
l = []
iden_status = 0
sleeper = 0
talash = 0

if not os.path.isfile('divar-car-{}.xlsx'.format(timenow)):
    workbook_car = Workbook()
    sh_car = workbook_car.active
    sh_car['A1'] = 'ID'
    sh_car['B1'] = 'Group'
    sh_car['C1'] = 'Title'; sh_car.column_dimensions['C'].width = 30
    sh_car['D1'] = 'Url'
    sh_car['E1'] = 'Location'
    sh_car['F1'] = 'Production year'
    sh_car['G1'] = 'Model'
    sh_car['H1'] = 'Kilometre'
    sh_car['I1'] = 'Color'
    sh_car['J1'] = 'Gearbox'
    sh_car['K1'] = 'Body'
    sh_car['L1'] = 'Description'; sh_car.column_dimensions['L'].width = 70
    sh_car['M1'] = 'Price'; sh_car.column_dimensions['M'].width = 15
    sh_car['N1'] = 'Date'
    sh_car['O1'] = 'Pictures'

else:
    workbook_car = load_workbook('divar-car-{}.xlsx'.format(timenow))
    sh_car = workbook_car.worksheets[0]
    for i in range(2, sh_car.max_row+1):
        id_list.append(sh_car.cell(row=i, column=1).value.strip())



if not os.path.isfile('divar-home-{}.xlsx'.format(timenow)):
    workbook_home = Workbook()
    sh_home = workbook_home.active
    sh_home['A1'] = 'ID'
    sh_home['B1'] = 'Group'
    sh_home['C1'] = 'Title'; sh_home.column_dimensions['C'].width = 30
    sh_home['D1'] = 'Url'
    sh_home['E1'] = 'Location'
    sh_home['F1'] = 'Production year'
    sh_home['G1'] = 'Area'
    sh_home['H1'] = 'Rooms'
    sh_home['I1'] = 'Total price'
    sh_home['J1'] = 'Price per meter'
    sh_home['K1'] = 'Deposit'
    sh_home['L1'] = 'Rent'
    sh_home['M1'] = 'Deposit_Rent'
    sh_home['N1'] = 'Advertiser'
    sh_home['O1'] = 'Floor'
    sh_home['P1'] = 'Elevator'
    sh_home['Q1'] = 'Parking'
    sh_home['R1'] = 'Warehousei'
    sh_home['S1'] = 'Description'; sh_home.column_dimensions['S'].width = 70
    sh_home['T1'] = 'Date'
    sh_home['U1'] = 'Pictures'

else:
    workbook_home = load_workbook('divar-home-{}.xlsx'.format(timenow))
    sh_home = workbook_home.worksheets[0]
    for i in range(2, sh_home.max_row+1):
        id_list.append(sh_home.cell(row=i, column=1).value.strip())




if not os.path.isfile('divar-motor-{}.xlsx'.format(timenow)):
    workbook_motor = Workbook()
    sh_motor = workbook_motor.active
    sh_motor['A1'] = 'ID'
    sh_motor['B1'] = 'Group'
    sh_motor['C1'] = 'Title'; sh_motor.column_dimensions['C'].width = 30
    sh_motor['D1'] = 'Url'
    sh_motor['E1'] = 'Location'
    sh_motor['F1'] = 'Production year'
    sh_motor['G1'] = 'Model'
    sh_motor['H1'] = 'Kilometre'
    sh_motor['I1'] = 'Description'; sh_motor.column_dimensions['I'].width = 70
    sh_motor['J1'] = 'Price'; sh_motor.column_dimensions['J'].width = 15
    sh_motor['K1'] = 'Date'
    sh_motor['L1'] = 'Pictures'

else:
    workbook_motor = load_workbook('divar-motor-{}.xlsx'.format(timenow))
    sh_motor = workbook_motor.worksheets[0]
    for i in range(2, sh_motor.max_row+1):
        id_list.append(sh_motor.cell(row=i, column=1).value.strip())


if not os.path.isfile('divar-{}.xlsx'.format(timenow)):
    workbook = Workbook()
    sh = workbook.active
    sh['A1'] = 'ID'
    sh['B1'] = 'Group'
    sh['C1'] = 'Title'; sh.column_dimensions['C'].width = 30
    sh['D1'] = 'Url'
    sh['E1'] = 'Location'
    sh['F1'] = 'Description'; sh.column_dimensions['F'].width = 70
    sh['G1'] = 'Price'; sh.column_dimensions['G'].width = 20
    sh['H1'] = 'Date'
    sh['I1'] = 'Pictures'

else:
    workbook = load_workbook('divar-{}.xlsx'.format(timenow))
    sh = workbook.worksheets[0]
    for i in range(2, sh.max_row+1):
        id_list.append(sh.cell(row=i, column=1).value.strip())




def runner():
    global talash
    global sleeper

    url = city_list.pop()
    r = requests.Session()
    r.headers = {'Connection': "close", "Accept": "*/*", "Content-type": "application/x-www-form-urlencoded; charset=UTF-8", "Accept-Language": "en-US", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    while 1:
        url_status_re = 0
        for i in range(1,51):
            if url_status_re == 1:
                break

            rr = r.get("{}/?page={}".format(url,str(i)))

            source_html = rr.text
            bs = BeautifulSoup(source_html, 'html.parser')
            urls_ = []

            for i in bs.find_all('a', {'class':"kt-post-card kt-post-card--outlined kt-post-card--bordered"}):
                urls_.append(i['href'])

            for lurl in urls_:
                id_ = lurl.split('/')[-1]

                if id_ in id_list:
                    url_status_re = 1
                    continue
                else:
                    id_list.append(id_)
                    url_status_re = 0
                
                verf = 1
                while verf:
                    try:
                        r_= r.get("http://divar.ir"+lurl)
                        verf = 0
                    except:
                        verf = 1

                html_source = r_.text
                bs_ = BeautifulSoup(html_source, 'html.parser')
                title = "|". join(bs_.find('title').text.split('|')[0:-1]).strip()
                images = bs_.find_all('img', {"class":'kt-image-block__image'})
                pictures = []

                for image in images:
                    pictures.append(image['src'])

                pictures = " , ".join(pictures)
                
                publish_time = datetime.datetime.now().strftime("%Y/%m/%d %H:%M")
                print(publish_time)
                group = ' '
                location = ' '
                type_ = ' '
                price = ' '
                zz = bs_.find_all("div", {'class':"post-info"})
                cc = []

                
                for i in zz:
                    cd = i.find_all("p", {'class':"kt-base-row__title kt-unexpandable-row__title"})
                    for j in cd:
                        cc.append(j)
                
                for i in bs_.find_all("div", {'class':'kt-base-row__start kt-unexpandable-row__title-box'}):
                    dd = i.find_all("p", {'class':'kt-base-row__title kt-unexpandable-row__title'})
                    for j in dd:
                        cc.append(j)
                
                for i in cc:
                    if i.text=='دسته‌بندی':
                        group = i.find_next().text.strip()
                    elif i.text=='محل':
                        location = i.find_next().text.strip()
                    elif i.text=='قیمت' or i.text=='قیمت کل':
                        price = i.find_next().text.strip()
                    elif i.text=='قیمت':
                        pass
                    else:
                        pass

                try:
                    description = zz[-1].find('p', {'class':"kt-description-row__text post-description kt-description-row__text--primary"}).text.strip()
                except Exception as error:
                    description = ' '
                
                berand = ' '
                year_ = ' '
                kilometre = ' '
                badane = ' '
                color = ' '
                gearbox = ' '
                forosh = ' '
                sanad = ' '
                if "سواری" == group or "سنگین" == group:
                    for i in cc:
                        if i.text == 'برند و مدل':
                            berand = i.find_next().text.strip()
                        elif i.text == 'سال ساخت':
                            year_ = i.find_next().text.strip()
                        elif i.text == 'کارکرد':
                            kilometre = i.find_next().text.strip()
                        elif i.text == 'وضعیت بدنه':
                            badane = i.find_next().text.strip()
                        elif i.text == 'رنگ':
                            color = i.find_next().text.strip()
                        elif i.text == 'گیربکس':
                            gearbox = i.find_next().text.strip()
                        elif i.text == 'نحوه فروش':
                            forosh = i.find_next().text.strip()
                        elif i.text == 'سند':
                            sanad = i.find_next().text.strip()
                        else:
                            pass

                    expens = [id_, group, title, "https://divar.ir"+lurl, location, year_, berand, kilometre, color, gearbox, badane, description, price, publish_time, pictures]
                    sh_car.append(expens)
                    workbook_car.save('divar-car-{}.xlsx'.format(timenow))

                elif 'موتورسیکلت' in group:
                    for i in cc:
                        if i.text == 'برند و مدل':
                            berand = i.find_next().text.strip()
                        elif i.text == 'سال ساخت':
                            year_ = i.find_next().text.strip()
                        elif i.text == 'کارکرد':
                            kilometre = i.find_next().text.strip()
                        else:
                            pass

                    expens = [id_, group, title, "https://divar.ir"+lurl, location, year_, berand, kilometre, '', '', '',description, price, publish_time, pictures]
                    sh_car.append(expens)
                    workbook_car.save('divar-car-{}.xlsx'.format(timenow))
                
                elif 'آپارتمان' in group or 'خانه و ویلا' in group:
                    area = ''
                    year_ = ''
                    rooms = ''
                    total_price = ''
                    meter_price = ''
                    vadie = ''
                    ejare = ''
                    vadie_ejare = ''
                    ad = ''
                    floor = ''
                    elevator = ''
                    parking = ''
                    warehouse = ''
                    for i in cc:
                        if  i.text == 'متراژ':
                            area = i.find_next().text.strip()
                        elif i.text == 'سال ساخت':
                            year_ = i.find_next().text.strip()
                        elif i.text == 'تعداد اتاق':
                            rooms = i.find_next().text.strip()
                        elif i.text == 'قیمت کل':
                            total_price = i.find_next().text.strip()
                        elif i.text == 'قیمت هر متر':
                            meter_price = i.find_next().text.strip()
                        elif i.text == 'ودیعه':
                            vadie = i.find_next().text.strip()
                        elif i.text == 'اجارهٔ ماهانه':
                            ejare = i.find_next().text.strip()
                        elif i.text == 'ودیعه و اجاره':
                            vadie_ejare = i.find_next().text.strip()
                        elif i.text == 'آگهی‌دهنده':
                            ad = i.find_next().text.strip()
                        elif i.text == 'طبقه':
                            floor = i.find_next().text.strip()
                        elif i.text == 'آسانسور':
                            elevator = i.find_next().text.strip()
                        elif i.text == 'پارکینگ':
                            parking = i.find_next().text.strip()
                        elif i.text == 'انباری':
                            warehouse = i.find_next().text.strip()
                        else:
                            pass
                    expens = [id_, group, title, "https://divar.ir"+lurl, location, year_, area, rooms, total_price,  meter_price, vadie, ejare, vadie_ejare, ad, floor, elevator, parking, warehouse,description, publish_time, pictures]
                    sh_home.append(expens)
                    workbook_home.save('divar-home-{}.xlsx'.format(timenow))

                elif 'مغازه و غرفه' in group or 'زمین و کلنگی' in group:
                    area = ''
                    year_ = ''
                    rooms = ''
                    total_price = ''
                    meter_price = ''
                    vadie = ''
                    ejare = ''
                    vadie_ejare = ''
                    ad = ''
                    floor = ''
                    elevator = ''
                    parking = ''
                    warehouse = ''
                    for i in cc:
                        if  i.text == 'متراژ':
                            area = i.find_next().text.strip()
                        elif i.text == 'سال ساخت':
                            year_ = i.find_next().text.strip()
                        elif i.text == 'تعداد اتاق':
                            rooms = i.find_next().text.strip()
                        elif i.text == 'قیمت کل':
                            total_price = i.find_next().text.strip()
                        elif i.text == 'قیمت هر متر':
                            meter_price = i.find_next().text.strip()
                        elif i.text == 'ودیعه':
                            vadie = i.find_next().text.strip()
                        elif i.text == 'اجارهٔ ماهانه':
                            ejare = i.find_next().text.strip()
                        elif i.text == 'ودیعه و اجاره':
                            vadie_ejare = i.find_next().text.strip()
                        elif i.text == 'آگهی‌دهنده':
                            ad = i.find_next().text.strip()
                        else:
                            pass
                    expens = [id_, group, title, "https://divar.ir"+lurl, location, year_, area, rooms, total_price,  meter_price, ad, floor, elevator, parking, warehouse,description, publish_time, pictures]
                    sh_home.append(expens)
                    workbook_home.save('divar-home-{}.xlsx'.format(timenow))
                else:
                    #time.sleep(0.25)
                    expens = [id_, group, title, "https://divar.ir"+lurl, location, description, price, publish_time, pictures]
                    sh.append(expens)
                    workbook.save('divar-{}.xlsx'.format(timenow))

        time.sleep(21600)


class myThread(threading.Thread):
	def __init__(self, threadID, name, counter, lock1, lock2, lock3):
		threading.Thread.__init__(self)
		self.threadID = threadID
		self.name = name
		self.counter = counter
		self.lock1 = lock1
		self.lock2 = lock2
		self.lock3 = lock3

	def run(self):
		runner()

if __name__ == "__main__":
    attack_threads = []
    try:
        for i in range(len(city_list)):
            attack_threads.append(myThread(i, "Thread-{}".format(i), i, lock1, lock2, lock3))
            attack_threads[i].start()

        for i in range(len(city_list)):
            attack_threads[i].join()
    except:
        workbook.save('divar-{}.xlsx'.format(timenow))
        workbook_motor.save('divar-motor-{}.xlsx'.format(timenow))
        workbook_car.save('divar-car-{}.xlsx'.format(timenow))
